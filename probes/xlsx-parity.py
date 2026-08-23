#!/usr/bin/env python3
"""Ground truth for xmux harness parity checks.

Reads real .xlsx files with openpyxl and emits JSON fixtures that the vitest
parity suite (`addin/src/excel/parity.test.ts`) replays through the pane's own
pure logic: A1 parsing, grid rendering, format classification. Excel at runtime
is the authority for live reads; this is the independent second opinion the two
are cross-checked against, so a drift in address arithmetic or format handling
shows up as a red test instead of a wrong answer in front of the user.

Excel semantics mirrored here:
- dates/times are stored as serials (days since 1899-12-30);
- COUNT counts numbers only, COUNTA counts non-empty cells;
- AVERAGE/SUM/MIN/MAX ignore text and blanks entirely.

Usage:
  python3 probes/xlsx-parity.py FILE.xlsx [FILE.xlsx ...] --out probes/parity-fixtures

Fixtures stay local (the corpus never enters the repo); add
`probes/parity-fixtures/` to .gitignore.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# Corporate workbooks written by non-Excel tools carry style attributes openpyxl's
# strict parser rejects (`applyNumberForm` instead of `applyNumberFormat`). The flag is
# irrelevant here, so the CellStyle constructor drops unknown keys instead of dying.
from openpyxl.styles.cell_style import CellStyle

_STYLE_ATTRS = set(getattr(CellStyle, "__attrs__", ()) )
_ORIGINAL_INIT = CellStyle.__init__


def _tolerant_init(self, **kwargs):
    return _ORIGINAL_INIT(self, **{k: v for k, v in kwargs.items() if k in _STYLE_ATTRS})


CellStyle.__init__ = _tolerant_init

EPOCH = datetime(1899, 12, 30)
MAX_ROWS = 300
MAX_COLUMNS = 60
MAX_SHEETS = 8

# Mirrors `isDerivableFormat` in addin/src/excel/format-profile.ts. Keep the two
# in lockstep: a disagreement is exactly what this lane exists to catch.
LITERALS = re.compile(r'"[^"]*"|\[[^\]]*\]')
SCALED = re.compile(r"[0#],{2,}")
SEMANTIC = re.compile(r"[ymdhs%]", re.IGNORECASE)
SIMPLE = re.compile(r"^[#0?,.\s\-+()]+$")


def is_derivable(fmt: str) -> bool:
    core = LITERALS.sub("", fmt.strip())
    if core == "" or core.lower() == "general":
        return True
    if SCALED.search(core):
        return False
    if SEMANTIC.search(core):
        return False
    return SIMPLE.match(core) is not None


def to_serial(value):
    """Everything Excel would hand JavaScript as a number becomes that number."""
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        delta = value - EPOCH
        return delta.days + delta.seconds / 86400 + delta.microseconds / 86400e6
    if isinstance(value, date):
        return (value - EPOCH.date()).days
    if isinstance(value, time):
        return (value.hour * 3600 + value.minute * 60 + value.second) / 86400
    return value


def column_aggregates(matrix):
    """COUNT/COUNTA/blank/sum/average/min/max per column, Excel-style."""
    out = []
    width = max((len(row) for row in matrix), default=0)
    for c in range(width):
        numbers = []
        filled = 0
        blank = 0
        for row in matrix:
            if c >= len(row):
                blank += 1
                continue
            value = row[c]
            if value is None or (isinstance(value, str) and value.strip() == ""):
                blank += 1
                continue
            filled += 1
            serial = to_serial(value)
            if isinstance(serial, (int, float)):
                numbers.append(float(serial))
        total = sum(numbers)
        count = len(numbers)
        out.append(
            {
                "column": c + 1,
                "count": count,
                "filled": filled,
                "blank": blank,
                "sum": total,
                "average": total / count if count else None,
                "min": min(numbers) if numbers else None,
                "max": max(numbers) if numbers else None,
            }
        )
    return out


def sheet_fixture(ws, ws_formulas=None):
    dim = ws.calculate_dimension()  # like "A1:F120"
    min_row, min_col = ws.min_row or 1, ws.min_column or 1
    rows = min((ws.max_row or 1) - min_row + 1, MAX_ROWS)
    cols = min((ws.max_column or 1) - min_col + 1, MAX_COLUMNS)

    values = []
    formats = []
    formulas = []
    for r in range(rows):
        value_row = []
        format_row = []
        formula_row = []
        for c in range(cols):
            cell = ws.cell(row=min_row + r, column=min_col + c)
            value_row.append(to_serial(cell.value))
            format_row.append(cell.number_format or "General")
            # The non-data_only pass carries the formula TEXT; None means a constant.
            if ws_formulas is not None:
                raw = ws_formulas.cell(row=min_row + r, column=min_col + c).value
                formula_row.append(raw if isinstance(raw, str) and raw.startswith("=") else None)
            else:
                formula_row.append(None)
        values.append(value_row)
        formats.append(format_row)
        formulas.append(formula_row)

    unique_formats = sorted({f for row in formats for f in row})
    return {
        "sheet": ws.title,
        "usedRange": dim,
        "anchor": {"top": min_row, "left": min_col},
        "rows": len(values),
        "cols": cols,
        "values": values,
        "formats": formats,
        "formulas": formulas,
        # Every distinct format string and whether the pane may treat the display
        # as derivable. The TS side must agree on every single one.
        "formats_classification": {f: is_derivable(f) for f in unique_formats},
        "aggregates": column_aggregates(values),
    }


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("files", nargs="+", help=".xlsx files (not .xlsb, not encrypted)")
    parser.add_argument("--out", default="probes/parity-fixtures")
    args = parser.parse_args()

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    for path in args.files:
        book = openpyxl.load_workbook(path, data_only=True)
        book_formulas = openpyxl.load_workbook(path, data_only=False)
        stem = Path(path).stem.replace(" ", "_")[:40]
        for index, ws in enumerate(book.worksheets[:MAX_SHEETS]):
            fixture = sheet_fixture(ws, book_formulas.worksheets[index])
            fixture["file"] = Path(path).name
            target = out_dir / f"{stem}__{index:02d}_{ws.title[:24]}.json".replace("/", "_")
            target.write_text(json.dumps(fixture, ensure_ascii=False))
            print(f"{target.name}: {fixture['rows']}x{fixture['cols']} · usedRange {fixture['usedRange']}")
        book.close()


if __name__ == "__main__":
    main()
